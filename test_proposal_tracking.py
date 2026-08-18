import datetime
import os
import tempfile
import unittest
from unittest.mock import Mock, patch

from openpyxl import Workbook, load_workbook

import pcs_proposal_web as web
from proposal_tracking_cutover_flags import (
    MASTER_FLAG,
    READ_FLAG,
    SHADOW_WRITE_FLAG,
    WRITE_FLAG,
    load_proposal_tracking_cutover_flags,
)
from proposal_tracking_store import (
    ProposalContactOrganizationRequired,
    ProposalTrackingStore,
)
from contact_store import ContactStoreError


PROPOSAL_ID = "11111111-1111-4111-8111-111111111111"
RELATIONSHIP_ID = "22222222-2222-4222-8222-222222222222"


CURRENT_TRACKER_HEADERS = [
    "Customer",
    "Contact",
    "Email Address",
    "Lead Generated",
    "Submitted By",
    "Estimate Dt",
    "Proposal Dt",
    "Follow-Up",
    "Status",
    "Estimated By",
    "Response",
]


class ProposalTrackingCutoverFlagTests(unittest.TestCase):
    def test_flags_are_disabled_by_default(self):
        flags = load_proposal_tracking_cutover_flags({})
        self.assertFalse(flags.master_enabled)
        self.assertFalse(flags.reads_enabled)
        self.assertFalse(flags.writes_enabled)
        self.assertTrue(flags.spreadsheet_reads_active)
        self.assertTrue(flags.spreadsheet_writes_active)

    def test_capability_flags_require_master_flag(self):
        flags = load_proposal_tracking_cutover_flags({
            READ_FLAG: "1",
            WRITE_FLAG: "1",
            SHADOW_WRITE_FLAG: "1",
        })
        self.assertFalse(flags.reads_enabled)
        self.assertFalse(flags.writes_enabled)
        self.assertFalse(flags.shadow_writes_enabled)


class ProposalTrackingWriteAuthorityTests(unittest.TestCase):
    def setUp(self):
        self.flags = load_proposal_tracking_cutover_flags({
            "PCS_PROPOSAL_STORAGE_MODE": "supabase_shadow",
        })

    def test_authoritative_database_failure_does_not_mutate_shadow(self):
        store = Mock()
        store.update_entries.side_effect = RuntimeError("database unavailable")
        with patch.object(
            web, "load_proposal_tracking_cutover_flags", return_value=self.flags
        ), patch.object(
            web, "get_proposal_tracking_store", return_value=store
        ), patch.object(
            web, "_update_proposal_tracker_missing_entries_spreadsheet"
        ) as spreadsheet_write:
            with self.assertRaisesRegex(RuntimeError, "database unavailable"):
                web.update_proposal_tracker_missing_entries([{"id": PROPOSAL_ID}])
        spreadsheet_write.assert_not_called()

    def test_shadow_failure_does_not_undo_authoritative_database_success(self):
        store = Mock()
        store.update_entries.return_value = 1
        with patch.object(
            web, "load_proposal_tracking_cutover_flags", return_value=self.flags
        ), patch.object(
            web, "get_proposal_tracking_store", return_value=store
        ), patch.object(
            web,
            "_update_proposal_tracker_missing_entries_spreadsheet",
            side_effect=RuntimeError("workbook locked"),
        ):
            updated = web.update_proposal_tracker_missing_entries(
                [{"id": PROPOSAL_ID}]
            )
        self.assertEqual(updated, 1)

    def test_shadow_mode_keeps_spreadsheet_writes_active(self):
        flags = load_proposal_tracking_cutover_flags({
            MASTER_FLAG: "true",
            WRITE_FLAG: "true",
            SHADOW_WRITE_FLAG: "true",
        })
        self.assertTrue(flags.writes_enabled)
        self.assertTrue(flags.spreadsheet_writes_active)
        self.assertFalse(flags.fully_cut_over)

    def test_full_cutover_disables_all_spreadsheet_access(self):
        flags = load_proposal_tracking_cutover_flags({
            MASTER_FLAG: "true",
            READ_FLAG: "true",
            WRITE_FLAG: "true",
            SHADOW_WRITE_FLAG: "false",
        })
        self.assertTrue(flags.fully_cut_over)
        self.assertFalse(flags.spreadsheet_reads_active)
        self.assertFalse(flags.spreadsheet_writes_active)

    def test_database_source_can_keep_spreadsheet_shadow_writes(self):
        flags = load_proposal_tracking_cutover_flags({
            "PCS_PROPOSAL_STORAGE_MODE": "supabase_shadow",
        })
        self.assertTrue(flags.database_source_authoritative)
        self.assertFalse(flags.spreadsheet_reads_active)
        self.assertTrue(flags.spreadsheet_writes_active)
        self.assertTrue(flags.shadow_writes_enabled)

    def test_persistent_shadow_mode_is_used_when_environment_is_absent(self):
        persisted = {
            MASTER_FLAG: "true",
            READ_FLAG: "false",
            WRITE_FLAG: "true",
            SHADOW_WRITE_FLAG: "true",
        }
        with patch(
            "pcs_local_settings.proposal_tracking_cutover_environment",
            return_value=persisted,
        ), patch.dict(os.environ, {}, clear=True):
            flags = load_proposal_tracking_cutover_flags()
        self.assertTrue(flags.master_enabled)
        self.assertFalse(flags.reads_enabled)
        self.assertTrue(flags.writes_enabled)
        self.assertTrue(flags.shadow_writes_enabled)
        self.assertTrue(flags.spreadsheet_reads_active)
        self.assertTrue(flags.spreadsheet_writes_active)

    def test_process_environment_overrides_persistent_cutover_setting(self):
        persisted = {
            MASTER_FLAG: "true",
            READ_FLAG: "false",
            WRITE_FLAG: "true",
            SHADOW_WRITE_FLAG: "true",
        }
        with patch(
            "pcs_local_settings.proposal_tracking_cutover_environment",
            return_value=persisted,
        ), patch.dict(os.environ, {MASTER_FLAG: "false"}, clear=True):
            flags = load_proposal_tracking_cutover_flags()
        self.assertFalse(flags.master_enabled)
        self.assertFalse(flags.writes_enabled)
        self.assertFalse(flags.shadow_writes_enabled)


class ProposalTrackingStoreTests(unittest.TestCase):
    def setUp(self):
        self.store = ProposalTrackingStore("https://example.supabase.co", "test-key")

    @staticmethod
    def proposal_row(**updates):
        row = {
            "proposal_id": PROPOSAL_ID,
            "lead_source": "Referral",
            "submitted_by": "David",
            "estimated_by": "Vern",
            "estimate_completed_date": None,
            "proposal_sent_date": "2026-07-15",
            "follow_up_date": None,
            "follow_up_required": True,
            "response_notes": None,
            "proposal": {
                "id": PROPOSAL_ID,
                "customer_name": "Example Roofing",
                "project_street_address": "123 Main St",
                "display_name": "Example Roofing - 123 Main St",
                "proposal_contact": [{
                    "is_primary": True,
                    "organization_contact": {
                        "id": RELATIONSHIP_ID,
                        "business_email": "casey@example.com",
                        "is_current": True,
                        "contact": {
                            "id": "contact-id",
                            "full_name": "Casey Smith",
                        },
                    },
                }],
            },
        }
        row.update(updates)
        return row

    def test_screen_entry_uses_generated_display_name_and_contact_tables(self):
        entry = self.store._screen_entry(self.proposal_row())
        self.assertEqual(entry["customer"], "Example Roofing - 123 Main St")
        self.assertEqual(entry["contact"], "Casey Smith")
        self.assertEqual(entry["email_address"], "casey@example.com")
        self.assertEqual(entry["lead_source"], "Referral")

    def test_management_list_is_rooted_in_proposal_and_filters_joined_status(self):
        rows = [{
            "id": PROPOSAL_ID,
            "customer_name": "Example Roofing",
            "project_street_address": "123 Main St",
            "display_name": "Example Roofing - 123 Main St",
            "proposal_folder_name": "Example Roofing - 123 Main St",
            "draft_detail": {"flat_roof_squares": "125"},
            "created_at": "2026-07-15T12:00:00+00:00",
            "updated_at": "2026-07-16T12:00:00+00:00",
            "proposal_tracking": {
                "status": "sent",
                "submitted_by": "Mark",
                "estimated_by": "Vern",
                "estimate_completed_date": "2026-07-14",
                "proposal_sent_date": "2026-07-15",
                "follow_up_date": "2026-07-29",
                "created_at": "2026-07-15T12:00:00+00:00",
                "updated_at": "2026-07-17T12:00:00+00:00",
            },
        }]
        with patch.object(self.store, "_request", return_value=rows) as request:
            entries = self.store.list_management_proposals(
                {"draft", "sent", "under_contract"}
            )
        self.assertEqual(request.call_args.args[0], "proposal")
        params = request.call_args.kwargs["params"]
        self.assertIn("proposal_tracking!inner", params["select"])
        self.assertEqual(
            params["proposal_tracking.status"],
            "in.(draft,sent,under_contract)",
        )
        self.assertEqual(entries[0]["status"], "sent")
        self.assertEqual(entries[0]["submitted_by"], "Mark")
        self.assertEqual(entries[0]["estimated_by"], "Vern")
        self.assertEqual(entries[0]["estimate_completed_date_display"], "7/14/2026")
        self.assertEqual(entries[0]["proposal_sent_date"], "2026-07-15")
        self.assertEqual(entries[0]["proposal_sent_date_display"], "7/15/2026")
        self.assertEqual(entries[0]["follow_up_date_display"], "7/29/2026")
        self.assertEqual(entries[0]["folder_name"], "Example Roofing - 123 Main St")
        self.assertEqual(entries[0]["last_modified_display"], "07/17/2026")

    def test_management_list_rejects_unknown_status_before_request(self):
        with patch.object(self.store, "_request") as request:
            with self.assertRaisesRegex(Exception, "Unsupported proposal status"):
                self.store.list_management_proposals({"open"})
        request.assert_not_called()

    def test_management_proposal_can_be_resolved_by_folder_name(self):
        row = {
            "id": PROPOSAL_ID,
            "customer_name": "Example Roofing",
            "project_street_address": "123 Main St",
            "display_name": "Example Roofing - 123 Main St",
            "proposal_folder_name": "Example Folder",
            "proposal_tracking": {"status": "draft"},
        }
        with patch.object(self.store, "_request", return_value=[row]) as request:
            proposal = self.store.get_management_proposal_by_folder(
                "Example Folder"
            )
        self.assertEqual(proposal["id"], PROPOSAL_ID)
        self.assertEqual(
            request.call_args.kwargs["params"]["proposal_folder_name"],
            "eq.Example Folder",
        )

    def test_management_entry_includes_primary_contact_fields(self):
        row = {
            "id": PROPOSAL_ID,
            "customer_name": "Example Roofing",
            "project_street_address": "123 Main St",
            "display_name": "Example Roofing - 123 Main St",
            "proposal_folder_name": "Example Roofing - 123 Main St",
            "draft_detail": {"flat_roof_squares": "125"},
            "proposal_tracking": {"status": "sent"},
            "proposal_contact": [{
                "organization_contact_id": RELATIONSHIP_ID,
                "is_primary": True,
                "contact_role": "primary",
                "organization_contact": {
                    "id": RELATIONSHIP_ID,
                    "business_email": "casey@example.com",
                    "contact": {"full_name": "Casey Smith"},
                    "organization": {"name": "Example Roofing"},
                },
            }],
        }
        entry = self.store._management_entry(row)
        self.assertEqual(entry["organization_contact_id"], RELATIONSHIP_ID)
        self.assertEqual(entry["contact_name"], "Casey Smith")
        self.assertEqual(entry["contact_email"], "casey@example.com")
        self.assertEqual(entry["organization_name"], "Example Roofing")
        self.assertEqual(entry["draft_detail"]["flat_roof_squares"], "125")

    def test_proposal_draft_detail_is_patched_by_proposal_id(self):
        detail = {"flat_roof_squares": "125", "product": "Gaco"}
        with patch.object(
            self.store, "_request", return_value=[{"id": PROPOSAL_ID}]
        ) as request:
            self.store.save_proposal_draft_detail(PROPOSAL_ID, detail)
        request.assert_called_once_with(
            "proposal",
            method="PATCH",
            params={"id": f"eq.{PROPOSAL_ID}"},
            payload={"draft_detail": detail},
            return_rows=True,
        )

    def test_management_contact_options_are_normalized_for_autocomplete(self):
        rows = [{
            "id": RELATIONSHIP_ID,
            "business_email": "casey@example.com",
            "is_current": True,
            "contact": {"id": "contact-id", "full_name": "Casey Smith"},
            "organization": {"id": "organization-id", "name": "Example Roofing"},
        }]
        with patch.object(self.store, "_request", return_value=rows) as request:
            options = self.store.list_management_contact_options()
        self.assertEqual(options, [{
            "id": RELATIONSHIP_ID,
            "name": "Casey Smith",
            "email": "casey@example.com",
            "organization": "Example Roofing",
        }])
        self.assertEqual(request.call_args.args[0], "organization_contact")
        self.assertEqual(request.call_args.kwargs["params"]["is_current"], "eq.true")

    def test_assign_existing_contact_sets_it_as_primary(self):
        relationship = {
            "id": RELATIONSHIP_ID,
            "business_email": "casey@example.com",
            "is_current": True,
            "contact": {"full_name": "Casey Smith"},
            "organization": {"name": "Example Roofing"},
        }
        responses = [[relationship], [{"id": PROPOSAL_ID}], [], []]
        with patch.object(self.store, "_request", side_effect=responses) as request:
            result = self.store.assign_or_create_primary_contact(
                PROPOSAL_ID,
                organization_contact_id=RELATIONSHIP_ID,
            )
        self.assertFalse(result["created"])
        self.assertEqual(result["name"], "Casey Smith")
        self.assertEqual(request.call_args.args[0], "proposal_contact")
        self.assertEqual(request.call_args.kwargs["method"], "POST")
        self.assertEqual(
            request.call_args.kwargs["payload"]["organization_contact_id"],
            RELATIONSHIP_ID,
        )

    def test_new_contact_requires_organization_when_domain_is_unknown(self):
        with patch.object(self.store, "_request", return_value=[]), patch.object(
            self.store, "find_organization_for_email", return_value=None
        ):
            with self.assertRaises(ProposalContactOrganizationRequired) as raised:
                self.store.assign_or_create_primary_contact(
                    PROPOSAL_ID,
                    contact_name="Casey Smith",
                    email="casey@new-roofer.com",
                )
        self.assertEqual(raised.exception.domain, "new-roofer.com")

    def test_new_contact_uses_prompted_organization_and_links_proposal(self):
        relationship = {
            "id": RELATIONSHIP_ID,
            "business_email": "casey@new-roofer.com",
            "is_current": True,
            "contact": {"full_name": "Casey Smith"},
            "organization": {"name": "New Roofer"},
        }
        responses = [[], [relationship], [{"id": PROPOSAL_ID}], [], []]
        with patch.object(self.store, "_request", side_effect=responses), patch.object(
            self.store, "find_organization_for_email", return_value=None
        ), patch.object(
            self.store,
            "resolve_named_organization_for_email",
            return_value="organization-id",
        ) as organization, patch.object(
            self.store, "create_contact", return_value="contact-id"
        ):
            result = self.store.assign_or_create_primary_contact(
                PROPOSAL_ID,
                contact_name="Casey Smith",
                email="casey@new-roofer.com",
                organization_name="New Roofer",
            )
        self.assertTrue(result["created"])
        self.assertEqual(result["organization"], "New Roofer")
        organization.assert_called_once_with("New Roofer", "casey@new-roofer.com")

    def test_customer_name_update_is_tenant_scoped_by_store_request(self):
        with patch.object(
            self.store,
            "_request",
            return_value=[{"id": PROPOSAL_ID}],
        ) as request:
            self.store.update_proposal_customer_name(
                PROPOSAL_ID,
                " Example   Management ",
            )

        request.assert_called_once_with(
            "proposal",
            method="PATCH",
            params={"id": f"eq.{PROPOSAL_ID}"},
            payload={"customer_name": "Example Management"},
            return_rows=True,
        )

    def test_missing_entries_include_missing_estimate_date(self):
        with patch.object(self.store, "list_proposals", return_value=[self.proposal_row()]):
            entries = self.store.list_missing_entries()
        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0]["estimate_date_input"], "")

    def test_completed_entry_is_not_returned_as_missing(self):
        row = self.proposal_row(estimate_completed_date="2026-07-14")
        with patch.object(self.store, "list_proposals", return_value=[row]):
            self.assertEqual(self.store.list_missing_entries(), [])

    def test_dead_entry_is_not_returned_as_missing(self):
        row = self.proposal_row(status="dead")
        with patch.object(self.store, "list_proposals", return_value=[row]):
            self.assertEqual(self.store.list_missing_entries(), [])

    def test_weekly_follow_up_queue_requires_explicit_eligibility(self):
        with patch.object(self.store, "_request", return_value=[]) as request:
            self.assertEqual(
                self.store.list_weekly_follow_ups(datetime.date(2026, 8, 3)),
                [],
            )
        params = request.call_args.kwargs["params"]
        self.assertEqual(params["status"], "eq.sent")
        self.assertEqual(params["proposal_sent_date"], "lte.2026-08-03")
        self.assertEqual(params["follow_up_date"], "is.null")
        self.assertEqual(params["follow_up_required"], "eq.true")

    def test_proposal_resolution_accepts_only_immutable_uuid(self):
        with patch.object(self.store, "_request") as request:
            self.assertEqual(self.store._resolve_proposal_id(PROPOSAL_ID), PROPOSAL_ID)
            self.assertEqual(self.store._resolve_proposal_id("42"), "")
            self.assertEqual(self.store._resolve_proposal_id("not-a-proposal-id"), "")
        request.assert_not_called()

    def test_mark_followups_updates_resolved_ids(self):
        with patch.object(
            self.store,
            "_request",
            return_value=[{"proposal_id": PROPOSAL_ID}],
        ) as request:
            count = self.store.mark_follow_ups(
                [PROPOSAL_ID], datetime.date(2026, 8, 3)
            )
        self.assertEqual(count, 1)
        self.assertEqual(
            request.call_args.kwargs["params"]["proposal_id"],
            f"in.({PROPOSAL_ID})",
        )
        self.assertEqual(request.call_args.kwargs["payload"], {
            "follow_up_date": "2026-08-03",
            "status": "sent",
        })

    def test_under_contract_status_aliases_are_normalized(self):
        for status in ("under contract", "under-contract", "under_contract"):
            with self.subTest(status=status):
                payload = self.store._editable_payload({"status": status})
                self.assertEqual(payload["status"], "under_contract")

    def test_editing_dates_does_not_reopen_a_closed_proposal(self):
        payload = self.store._editable_payload({
            "proposal_date": "8/1/2026",
            "follow_up_date": "8/3/2026",
        })
        self.assertNotIn("status", payload)

    def test_legacy_statuses_map_to_new_lifecycle(self):
        self.assertEqual(
            self.store._editable_payload({"status": "won"})["status"],
            "under_contract",
        )
        self.assertEqual(
            self.store._editable_payload({"status": "withdrawn"})["status"],
            "dead",
        )

    def test_finished_is_an_accepted_lifecycle_status(self):
        self.assertEqual(
            self.store._editable_payload({"status": "finished"})["status"],
            "finished",
        )

    def test_new_proposal_save_sets_estimate_date_without_setting_sent_date(self):
        responses = [[], [], [{"id": PROPOSAL_ID}], []]
        with patch.object(self.store, "_request", side_effect=responses) as request:
            proposal_id = self.store.upsert_from_proposal_save(
                created_date="08/03/2026",
                customer_name="Example Roofing",
                street_address="123 Main St",
                city="Denver",
                state="co",
                zip_code="80202",
                submitted_by="David",
                folder_name="Example Roofing - 123 Main St",
                lead_value="Referral",
            )
        self.assertEqual(proposal_id, PROPOSAL_ID)
        proposal_payload = request.call_args_list[2].kwargs["payload"]
        tracking_payload = request.call_args_list[3].kwargs["payload"]
        self.assertEqual(proposal_payload["project_state"], "CO")
        self.assertNotIn("estimate_completed_date", proposal_payload)
        self.assertEqual(tracking_payload["proposal_id"], PROPOSAL_ID)
        self.assertEqual(
            tracking_payload["estimate_completed_date"], "2026-08-03"
        )
        self.assertNotIn("proposal_sent_date", tracking_payload)

    def test_failed_tracking_insert_removes_new_proposal_identity(self):
        responses = [
            [],
            [],
            [{"id": PROPOSAL_ID}],
            ContactStoreError("Tracking insert failed"),
            [],
        ]
        with patch.object(self.store, "_request", side_effect=responses) as request:
            with self.assertRaisesRegex(ContactStoreError, "Tracking insert failed"):
                self.store.upsert_from_proposal_save(
                    created_date=None,
                    customer_name="Example Roofing",
                    street_address="123 Main St",
                    city="Denver",
                    state="CO",
                    zip_code="80202",
                    submitted_by="Mark",
                    folder_name="Example Roofing - 123 Main St",
                    estimated_by="",
                )
        cleanup = request.call_args_list[-1]
        self.assertEqual(cleanup.args[0], "proposal")
        self.assertEqual(cleanup.kwargs["method"], "DELETE")
        self.assertEqual(cleanup.kwargs["params"]["id"], f"eq.{PROPOSAL_ID}")

    def test_existing_contact_draft_is_finalized_by_id(self):
        responses = [
            [{"id": PROPOSAL_ID}],
            [],
            [{"proposal_id": PROPOSAL_ID}],
            [],
        ]
        with patch.object(self.store, "_request", side_effect=responses) as request:
            proposal_id = self.store.upsert_from_proposal_save(
                proposal_id=PROPOSAL_ID,
                created_date="08/13/2026",
                customer_name="Boulder County",
                street_address="132 Main St",
                city="Denver",
                state="CO",
                zip_code="88888",
                submitted_by="Vern",
                folder_name="Boulder County - 132 Main St",
            )

        self.assertEqual(proposal_id, PROPOSAL_ID)
        lookup = request.call_args_list[0]
        self.assertEqual(lookup.kwargs["params"]["id"], f"eq.{PROPOSAL_ID}")
        proposal_update = request.call_args_list[1]
        self.assertEqual(
            proposal_update.kwargs["payload"]["proposal_folder_name"],
            "Boulder County - 132 Main St",
        )
        tracking_update = request.call_args_list[3]
        self.assertEqual(
            tracking_update.kwargs["payload"]["estimate_completed_date"],
            "2026-08-13",
        )


class ProposalTrackingSpreadsheetColumnTests(unittest.TestCase):
    def setUp(self):
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.tracker_path = os.path.join(
            self.temporary_directory.name, "Proposal Tracking.xlsx"
        )

    def tearDown(self):
        self.temporary_directory.cleanup()

    def write_tracker(self, rows):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Tracking"
        worksheet.append(CURRENT_TRACKER_HEADERS)
        for row in rows:
            worksheet.append(row)
        workbook.save(self.tracker_path)
        workbook.close()

    def read_row(self, row_number=2):
        workbook = load_workbook(self.tracker_path, data_only=True)
        try:
            worksheet = workbook.active
            return [
                worksheet.cell(row=row_number, column=column).value
                for column in range(1, 12)
            ]
        finally:
            workbook.close()

    def test_tracker_screen_reads_reordered_columns_by_header(self):
        self.write_tracker([[
            "Anchor Roofing - 10 Oak Ave",
            "",
            "joel.anchorroofing@gmail.com",
            "Referral",
            "Mark",
            datetime.date(2026, 7, 30),
            "-",
            "-",
            "Draft Unsent",
            "Mark",
            "Waiting",
        ]])

        entries, error = web._load_proposal_tracker_missing_entries_spreadsheet(
            self.tracker_path
        )

        self.assertIsNone(error)
        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0]["estimated_by"], "Mark")
        self.assertEqual(entries[0]["estimate_date_input"], "7/30/2026")
        self.assertEqual(entries[0]["proposal_date_input"], "-")
        self.assertEqual(entries[0]["follow_up_date_input"], "-")
        self.assertEqual(entries[0]["status"], "draft")

    def test_legacy_tracker_gets_status_column_without_overwriting_response(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append([
            "Customer", "Contact", "Email Address", "Lead Generated", "Submitted By",
            "Estimate Dt", "Proposal Dt", "Follow-Up", "Estimated By", "Response",
        ])
        worksheet.append([
            "Example Roofing - 123 Main St", "Casey", "casey@example.com", "Referral",
            "David", "", "", "", "Vern", "Keep this response",
        ])

        web._ensure_proposal_tracker_status_column(worksheet)

        self.assertEqual(worksheet.cell(1, 9).value, "Status")
        self.assertEqual(worksheet.cell(2, 10).value, "Vern")
        self.assertEqual(worksheet.cell(2, 11).value, "Keep this response")

    def test_dead_tracker_row_is_not_shown_on_tracker_screen(self):
        self.write_tracker([[
            "Closed Roofing - 500 Main St",
            "",
            "",
            "Referral",
            "David",
            "",
            "-",
            "-",
            "Dead",
            "Vern",
            "Closed",
        ]])

        entries, error = web._load_proposal_tracker_missing_entries_spreadsheet(
            self.tracker_path
        )

        self.assertIsNone(error)
        self.assertEqual(entries, [])

    def test_tracker_screen_save_writes_reordered_columns_by_header(self):
        self.write_tracker([[
            "Example Roofing - 123 Main St",
            "Casey",
            "casey@example.com",
            "Referral",
            "David",
            "",
            "",
            "",
            "Sent",
            "Vern",
            "Keep this response",
        ]])

        count = web._update_proposal_tracker_missing_entries_spreadsheet([{
            "row_number": "2",
            "contact": "Casey Smith",
            "email_address": "casey@example.com",
            "lead_source": "Website",
            "submitted_by": "Mark",
            "estimate_date": "7/30/2026",
            "proposal_date": "8/1/2026",
            "follow_up_date": "8/15/2026",
            "status": "under_contract",
            "estimated_by": "Mark",
        }], self.tracker_path)

        self.assertEqual(count, 1)
        row = self.read_row()
        self.assertEqual(row[5], "7/30/2026")
        self.assertEqual(row[6], "8/1/2026")
        self.assertEqual(row[7], "8/15/2026")
        self.assertEqual(row[8], "Under Contract")
        self.assertEqual(row[9], "Mark")
        self.assertEqual(row[10], "Keep this response")

    def test_weekly_follow_up_uses_proposal_and_follow_up_headers(self):
        self.write_tracker([[
            "Example Roofing - 123 Main St",
            "Casey",
            "casey@example.com",
            "Referral",
            "David",
            datetime.date(2026, 7, 30),
            datetime.date(2026, 8, 1),
            "",
            "Sent",
            "Vern",
            "Keep this response",
        ]])

        entries, _, error = web._load_weekly_follow_up_entries_spreadsheet(
            self.tracker_path, datetime.date(2026, 8, 2)
        )
        self.assertIsNone(error)
        self.assertEqual([entry["row_number"] for entry in entries], [2])

        count = web._update_weekly_follow_up_dates_spreadsheet(
            [2], datetime.date(2026, 8, 3), self.tracker_path
        )
        self.assertEqual(count, 1)
        row = self.read_row()
        self.assertEqual(row[6], datetime.datetime(2026, 8, 1, 0, 0))
        self.assertEqual(row[7], datetime.datetime(2026, 8, 3, 0, 0))
        self.assertEqual(row[8], "Sent")
        self.assertEqual(row[9], "Vern")
        self.assertEqual(row[10], "Keep this response")

    def test_proposal_save_refresh_preserves_dates_and_response_columns(self):
        self.write_tracker([[
            "Example Roofing - 123 Main St",
            "Casey",
            "casey@example.com",
            "Referral",
            "David",
            "",
            datetime.date(2026, 8, 1),
            "",
            "Sent",
            "Mark",
            "Keep this response",
        ]])

        updated = web.update_existing_tracker_row(
            "Example Roofing - 123 Main St",
            "Website",
            "David",
            datetime.date(2026, 7, 30),
            self.tracker_path,
        )

        self.assertTrue(updated)
        row = self.read_row()
        self.assertEqual(row[5], datetime.datetime(2026, 7, 30, 0, 0))
        self.assertEqual(row[6], datetime.datetime(2026, 8, 1, 0, 0))
        self.assertEqual(row[8], "Sent")
        self.assertEqual(row[9], "Vern")
        self.assertEqual(row[10], "Keep this response")

    def test_dead_status_is_excluded_from_weekly_follow_up(self):
        self.write_tracker([[
            "Example Roofing - 123 Main St",
            "Casey",
            "casey@example.com",
            "Referral",
            "David",
            datetime.date(2026, 7, 30),
            datetime.date(2026, 8, 1),
            "",
            "Dead",
            "Vern",
            "Closed",
        ]])

        entries, _, error = web._load_weekly_follow_up_entries_spreadsheet(
            self.tracker_path, datetime.date(2026, 8, 20)
        )

        self.assertIsNone(error)
        self.assertEqual(entries, [])


if __name__ == "__main__":
    unittest.main()
